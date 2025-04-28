from fastapi import FastAPI, HTTPException
from fastapi.responses import JSONResponse
from pydantic import BaseModel
import uvicorn
import base64
from PIL import Image
import io
import json
import logging
import os
from datetime import datetime
from typing import Optional
from contextlib import asynccontextmanager
import subprocess
import sys
from openpyxl import load_workbook, Workbook
import os


def update_excel_file(json_data: dict, ref_number: str):
    excel_file = 'business_cards.xlsx'
    
    headers = [
        'Image','Company','First Name','Last Name','Title','Email', 'Fax Number',
        'Website','Mobile Phone','Phone','Address','Street','State','Country',
        'Postal Code'
    ]
    
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Business Cards'
        ws.append(headers)

    key_mapping = {
        'company_name': 'Company',
        'first_name': 'First Name',
        'last_name': 'Last Name',
        'job_title': 'Title',
        'email_address': 'Email',
        'fax_detail': 'Fax Number',
        'website_link': 'Website',
        'mobile_phone': 'Mobile Phone',
        'phone': 'Phone',
        'complete_address': 'Address',
        'street': 'Street',
        'state': 'State',
        'country': 'Country',
        'postal_code': 'Postal Code'
    }

    row_data = [ref_number]  # Start with reference number as Image column
    for header in headers[1:]:  # Skip Image column in iteration
        value = 'NA'
        for json_key, excel_header in key_mapping.items():
            if excel_header == header and json_key in json_data:
                value = json_data[json_key]
                break
        row_data.append(value)
    
    ws.append(row_data)
    wb.save(excel_file)
# Add to the existing process_card endpoint after JSON processing:

def get_next_reference_number():
    counter_file = 'reference_counter.txt'
    try:
        with open(counter_file, 'r') as f:
            counter = int(f.read().strip())
    except FileNotFoundError:
        counter = 1
    
    with open(counter_file, 'w') as f:
        f.write(str(counter + 1))
    
    return f"REF{counter:06d}"
# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Pydantic models
class CardImage(BaseModel):
    image: str

class CardData(BaseModel):
    company_name: str
    first_name: str
    last_name: str
    title: str
    email_address: str
    fax_detail: str
    website_link : str
    mobile_phone: str
    phone: str
    complete_address: str
    street: str
    state: str
    country: str
    postal_code: str

class ProcessResponse(BaseModel):
    status: str
    data: Optional[dict] = None
    message: Optional[str] = None
    saved_image: Optional[str] = None

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Lifecycle event handler"""
    # Startup
    os.makedirs('saved_cards', exist_ok=True)
    logger.info("Server started successfully")
    yield
    # Shutdown
    logger.info("Server shutting down")

# Create FastAPI app with lifespan
app = FastAPI(
    title="Business Card Processor",
    description="API for processing business card images and extracting information",
    version="1.0.0",
    lifespan=lifespan
)

def run_ollama_command(image_path: str) -> str:
    """Run Ollama command and return output"""
    try:
        # Updated prompt that matches the data structure
        prompt1 = """
    You are an expert OCR AI specializing in extracting structured data from business cards, even with complex layouts.
    Your task is to accurately extract information while strictly adhering to the following JSON format.

    Instructions: 
    
    1. Start by extracting all the text on the card. 
    2. Classify all the data into the respective fields given in the JSON format accurately.
    3. If there is not data related to a particular field just give out an empty string " ".  
    
    
    JSON Format:

    {
        "company_name": "string",
        "first_name": "string",
        "last_name": "string",
        "job_title": "string",
        "email_address": "string",
        "complete_address": "string",
        "street": "string",
        "state": "string",
        "country": "string",
        "postal_code": "string",
        "fax_detail": "string",
        "mobile_phone": "string",
        "phone": "string",
        "website_link": "string"
    }
    Note: add direct contact to phone. 
    Notice that email, first name and last name exists in all the cards. Make sure to identify these.
    """
          
        # Convert to absolute path
        abs_path = os.path.abspath(image_path)
        cmd1 = f'OLLAMA_TEMPERATURE=0.0 OLLAMA_CTX_SIZE=8192 ollama run llama3.2-vision:latest "{prompt1}" "{abs_path}" --format json'
        
    
        logger.info(f"Running Ollama command: {cmd1}")
        
        # Run command with increased timeout
        process = subprocess.run(cmd1, shell=True,
            capture_output=True,
            text=True,
            timeout=180, encoding='utf-8')
        
        if process.returncode != 0:
            logger.error(f"Command failed with error: {process.stderr}")
            raise Exception(f"Ollama command failed: {process.stderr}")
            
        logger.info(f"Ollama response: {process.stdout}")
        return process.stdout
        
    except subprocess.TimeoutExpired:
        logger.error("Ollama command timed out after 180 seconds")
        raise Exception("Ollama processing timed out after 180 seconds")
    except Exception as e:
        logger.error(f"Error in Ollama command: {str(e)}")
        raise

## Process a business card image and extract information
@app.post("/process-card", response_model=ProcessResponse)
async def process_card(card_request: CardImage):
    try:
        # Decode and save image
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ref_number = get_next_reference_number()

        try:
            image_data = base64.b64decode(card_request.image)
            image = Image.open(io.BytesIO(image_data))
            logger.info(f"Received image: format={image.format}, size={image.size}, mode={image.mode}")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid image data: {str(e)}")

        # Save original image with absolute path
        save_dir = os.path.abspath('saved_cards')
        image_path = os.path.join(save_dir, f"{ref_number}.jpg")
        image.save(image_path, format='JPEG', quality=95)
        logger.info(f"Saved image to {image_path}")

        try:
            # Process with Ollama
            output = run_ollama_command(image_path)
            logger.info(f"Processing Ollama output: {output}")
            
            try:
                # Parse JSON from Ollama response
                parsed_data = json.loads(output)
                logger.info(f"Parsed JSON data: {parsed_data}")
                
                # Load existing JSON data or create new list
                try:
                    with open('business_cards.json', 'r', encoding='utf-8') as f:
                        try:
                            existing_data = json.load(f)
                            if not isinstance(existing_data, list):
                                existing_data = [existing_data] if existing_data else []
                        except json.JSONDecodeError:
                            existing_data = []
                except FileNotFoundError:
                    existing_data = []
                
                # Append new data to JSON list
                existing_data.append(parsed_data)
                
                # Write back the entire JSON list
                with open('business_cards.json', 'w', encoding='utf-8') as f:
                    json.dump(existing_data, f, indent=2, ensure_ascii=False)
                
                # Update Excel file with new data
                try:
                    update_excel_file(parsed_data, ref_number)
                except Exception as excel_error:
                    logger.error(f"Error updating Excel file: {str(excel_error)}")
                    return ProcessResponse(status="partial_success", data=parsed_data, message=f"Data saved to JSON but Excel update failed: {str(excel_error)}", saved_image=image_path)
                
                return ProcessResponse(status="success",data=parsed_data, saved_image=image_path)
                
            except json.JSONDecodeError as e:
                logger.error(f"Failed to parse JSON response: {e}")
                return ProcessResponse(status="error", message=f"Failed to parse response: {str(e)}", saved_image=image_path)

        except Exception as e:
            logger.error(f"Error processing with Ollama: {str(e)}")
            return ProcessResponse(status="partial_success", message=f"Image saved but processing failed: {str(e)}", saved_image=image_path)

    except Exception as e:
        logger.error(f"Error in process_card: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

## Save the extracted card data to a JSON file
def save_to_json(card_data: CardData):
    try:
        filename = 'business_cards.json'
        # Create file if it doesn't exist
        if not os.path.exists(filename):
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump([], f)
        
        ## Read existing data
        with open(filename, 'r', encoding='utf-8') as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError:
                data = []
        ## Append new data
        data.append(card_data.dict())
        # Write back to file
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)            
        logger.info(f"Saved card data to {filename}")
    except Exception as e:
        logger.error(f"Error saving to JSON: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error saving data: {str(e)}")

## Health check endpoint
@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "ollama_available": check_ollama_available()
    }

## Check if Ollama is available
def check_ollama_available() -> bool:
    try:
        subprocess.run(["ollama", "list"], capture_output=True, timeout=5)
        return True
    except:
        return False
        
## Start up uvicorn
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("shot:app", host="0.0.0.0", port=5000, reload=True)
